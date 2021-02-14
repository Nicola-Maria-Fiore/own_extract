from openpyxl import load_workbook
import xlsxwriter
import keyboard
import time
import sys
import os

out = "results/Excel Empty/"
out_original = "results/Excel Original/"

doImport_output = "results/Do Files/import.do"
doAppend_output = "results/Do Files/append.do"

company = "resources/list of firms.csv"
year = "resources/period.csv"
conf = "resources/conf.txt"

excel_path = None
times = []
formula = None

def createXslc(comapny_id, year_list):
    for o in [out,out_original]:
        workbook = xlsxwriter.Workbook(o+comapny_id+'.xlsx')

        for y in year_list:
            worksheet = workbook.add_worksheet(y)
            content = formula.replace("INSERT",comapny_id).replace("YEAR",y)
            worksheet.write('A1', content)
        
        workbook.close()

def openFiles(companies):
    for c in companies:
        f_path = out_original + c + ".xlsx"
        f_path = os.path.abspath(f_path)
        os.system('start "{}" "{}"'.format(excel_path,f_path))
        time.sleep(times[0])
        keyboard.send('ctrl+s')
        time.sleep(times[1])
        keyboard.send('alt+F4')
        time.sleep(times[2])

def readFile(fname):
    result = []
    with open(fname, 'r', encoding='utf-8') as file_in:
        for line in file_in:
            result.append(line.strip().replace('\n',""))
    return result

def getConf(s):
    return ((s.split("-->"))[1]).strip()

def createDo(companies, years):
    importDo = ""
    appendDo = ""
    first = True
    for c in companies:
        for y in years:
            importDo += 'import "{}.xlsx", sheet("{}") cellrange(A2) firstrow \n save "{}_{}.dta" \n'.format(c,y,c,y)
            if first: 
                appendDo += 'use "{}_{}.dta" \n'.format(c,y)
                first = False
            else:
                appendDo += 'append "{}_{}.dta" \n'.format(c,y)

    with open(doImport_output, 'w') as f:
        f.write(importDo)
    
    with open(doAppend_output, 'w') as f:
        f.write(appendDo)
    

def checkFiles(companies):
    for c in companies:
        f_path = out_original + c + ".xlsx"
        xlsx_file = load_workbook(f_path)
        sheet = xlsx_file.active
        value = sheet.cell(row=2, column=2).value 
        if value==None or len(value)==0:
            print(f_path)


if __name__ == "__main__":
    companies = readFile(company)
    companies = [c for c in companies if c!="#N/A"]
    years = readFile(year)

    configurations = readFile(conf)
    excel_path = getConf(configurations[0])
    formula = getConf(configurations[4])
    times = configurations[1:4]
    for i in range(0, len(times)):
        times[i] = int(getConf(times[i]))  

    if len(sys.argv)>1:
        if sys.argv[1] == "-a": 
            for c in companies:
                createXslc(c,years) 
        elif sys.argv[1] == "-b":     
            openFiles(companies)
        elif sys.argv[1] == "-c": 
            createDo(companies, years)
        elif sys.argv[1] == "-d": 
            checkFiles(companies)
        else:
            print("Error: Read instructions")
    else:
        print("Error: Read instructions") 
   
    print("Done!")